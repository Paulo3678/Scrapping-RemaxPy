import openpyxl

# CRIAR UMA PLANILHA(BOOK)
book = openpyxl.Workbook()

# COMO VIZUALIZAR PÁGINAS EXISTENTES
print(book.sheetnames)

# COMO CRIAR UMA PÁGINA
book.create_sheet('Frutas')

# SELECIONANDO UMA PÁGINA
frutas_page = book['Frutas']
# NOME DA LINHA
frutas_page.append(['Fruta', 'Quantidade', 'Preco'])


# ADICIONANDO DADO EM UM PÁGINA
frutas_page.append(["Banana", '5', 'R$3,90'])
frutas_page.append(["Uva", '3', 'R$5,90'])
frutas_page.append(["Abacate", '2', 'R$7,90'])

# SALVAR A PLANILHA
book.save('Planilha de compras.xlsx')